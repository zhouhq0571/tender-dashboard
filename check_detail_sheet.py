import openpyxl

# 读取银行旧关键词文件，检查详情sheet结构
wb = openpyxl.load_workbook('/Users/tmp/银行-20260709使用旧关键词.xlsx', data_only=False)
sheet_names = wb.sheetnames

print(f"Sheet总数: {len(sheet_names)}")
print(f"Sheet列表: {sheet_names[:10]}...")

# 检查第一个详情sheet（跳过招标基本信息）
if len(sheet_names) > 1:
    detail_sheet = wb[sheet_names[1]]
    print(f"\n详情sheet '{sheet_names[1]}' 结构:")
    print(f"  行数: {detail_sheet.max_row}, 列数: {detail_sheet.max_column}")
    
    # 读取标题行
    headers = [cell.value for cell in detail_sheet[1]]
    print(f"  标题行: {headers}")
    
    # 读取前3行数据
    print(f"\n  前3行数据:")
    for row in detail_sheet.iter_rows(min_row=2, max_row=min(4, detail_sheet.max_row), values_only=False):
        for col_idx, cell in enumerate(row):
            if cell.value:
                print(f"    列{col_idx}({headers[col_idx]}): {str(cell.value)[:100]}...")
        print()

wb.close()
