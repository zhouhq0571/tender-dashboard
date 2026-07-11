import openpyxl

# 读取银行旧关键词文件，检查详情sheet完整结构
wb = openpyxl.load_workbook('/Users/tmp/银行-20260709使用旧关键词.xlsx', data_only=False)
sheet_names = wb.sheetnames

# 检查第一个详情sheet（跳过招标基本信息）
if len(sheet_names) > 1:
    detail_sheet = wb[sheet_names[1]]
    print(f"详情sheet '{sheet_names[1]}' 完整数据:")
    print(f"行数: {detail_sheet.max_row}, 列数: {detail_sheet.max_column}")
    print()
    
    # 读取所有行
    for row in detail_sheet.iter_rows(min_row=1, max_row=detail_sheet.max_row, values_only=False):
        col0 = str(row[0].value) if row[0].value else ''
        col1 = str(row[1].value) if row[1].value else ''
        print(f"{col0}: {col1[:200]}")

wb.close()
