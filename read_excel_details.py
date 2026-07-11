import openpyxl
import json
from datetime import datetime

# 读取三个Excel文件
files = [
    ('信托', '/Users/tmp/信托-20260709.xlsx'),
    ('银行新关键词', '/Users/tmp/银行-20260709使用新关键词.xlsx'),
    ('银行旧关键词', '/Users/tmp/银行-20260709使用旧关键词.xlsx'),
]

all_projects = []

for label, filepath in files:
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet_names = wb.sheetnames
    
    # 读取第一个sheet（招标基本信息）
    ws_basic = wb[sheet_names[0]]
    headers = [cell.value for cell in ws_basic[1]]
    
    # 获取详情链接列索引
    detail_link_idx = headers.index('详情链接') if '详情链接' in headers else -1
    title_idx = headers.index('标题') if '标题' in headers else 0
    pub_date_idx = headers.index('发布时间') if '发布时间' in headers else 1
    company_idx = headers.index('招标单位名称') if '招标单位名称' in headers else 6
    contact_idx = headers.index('招标单位联系人') if '招标单位联系人' in headers else 7
    phone_idx = headers.index('招标单位联系电话') if '招标单位联系电话' in headers else 8
    province_idx = headers.index('省份') if '省份' in headers else 2
    city_idx = headers.index('城市') if '城市' in headers else 3
    preview_idx = headers.index('内容预览') if '内容预览' in headers else -1
    
    for row in ws_basic.iter_rows(min_row=2, values_only=False):
        # 获取详情链接URL
        bid_url = None
        bid_num = None
        if detail_link_idx >= 0 and row[detail_link_idx].hyperlink:
            bid_url = row[detail_link_idx].hyperlink.target
            # 从URL提取bid号
            if bid_url and 'qianlima.com' in bid_url:
                bid_num = bid_url.split('bid-')[-1].split('.')[0]
        
        # 获取基本信息
        title = str(row[title_idx].value) if row[title_idx].value else ''
        company = str(row[company_idx].value) if row[company_idx].value else ''
        pub_date = str(row[pub_date_idx].value) if row[pub_date_idx].value else ''
        contact = str(row[contact_idx].value) if row[contact_idx].value and row[contact_idx].value != 'nan' else '未披露'
        phone = str(row[phone_idx].value) if row[phone_idx].value and row[phone_idx].value != 'nan' else '未披露'
        province = str(row[province_idx].value) if row[province_idx].value else ''
        city = str(row[city_idx].value) if row[city_idx].value else ''
        preview = str(row[preview_idx].value) if preview_idx >= 0 and row[preview_idx].value else ''
        
        # 读取详情sheet中的项目正文
        detail_content = ''
        if bid_num and bid_num in wb.sheetnames:
            ws_detail = wb[bid_num]
            detail_headers = [cell.value for cell in ws_detail[1]]
            # 查找"项目正文"列
            content_idx = -1
            for i, h in enumerate(detail_headers):
                if h and '项目正文' in str(h):
                    content_idx = i
                    break
            
            if content_idx >= 0:
                for detail_row in ws_detail.iter_rows(min_row=2, values_only=False):
                    if detail_row[content_idx].value:
                        detail_content = str(detail_row[content_idx].value)
                        break
        
        all_projects.append({
            'source_file': label,
            'bid_num': bid_num,
            'bid_url': bid_url,
            'title': title,
            'company': company,
            'pub_date': pub_date,
            'contact': contact,
            'phone': phone,
            'province': province,
            'city': city,
            'preview': preview[:200],
            'detail_content': detail_content[:500] if detail_content else ''
        })
    
    wb.close()

# 保存结果
with open('/Users/zhouhq/Documents/kimi/workspace/bidding-daily/excel_all_projects.json', 'w', encoding='utf-8') as f:
    json.dump(all_projects, f, ensure_ascii=False, indent=2)

print(f"总计读取 {len(all_projects)} 个项目")
print(f"有详情sheet的项目: {sum(1 for p in all_projects if p['detail_content'])}")
print(f"有详情URL的项目: {sum(1 for p in all_projects if p['bid_url'])}")

# 显示前3个有详情内容的项目
print("\n--- 有详情内容的项目示例 ---")
for p in all_projects[:3]:
    if p['detail_content']:
        print(f"\n{p['company']} - {p['title'][:50]}")
        print(f"  详情URL: {p['bid_url']}")
        print(f"  项目正文: {p['detail_content'][:300]}...")
