import openpyxl
import json

# 读取三个Excel文件
files = [
    ('信托', '/Users/tmp/信托-20260709.xlsx'),
    ('银行新关键词', '/Users/tmp/银行-20260709使用新关键词.xlsx'),
    ('银行旧关键词', '/Users/tmp/银行-20260709使用旧关键词.xlsx'),
]

all_projects = []

for label, filepath in files:
    print(f"\n{'='*60}")
    print(f"文件: {label} ({filepath})")
    print(f"{'='*60}")
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet_names = wb.sheetnames
    print(f"Sheet列表: {sheet_names}")
    
    # 读取第一个sheet（招标基本信息）
    ws = wb[sheet_names[0]]
    print(f"\n第一个sheet名称: {ws.title}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 读取标题行
    headers = [cell.value for cell in ws[1]]
    print(f"\n标题行: {headers}")
    
    # 检查前3行的超链接
    print(f"\n--- 前3行数据（含超链接检查）---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(4, ws.max_row), values_only=False), 2):
        print(f"\n行 {row_idx}:")
        for col_idx, cell in enumerate(row):
            if cell.value:
                has_link = hasattr(cell, 'hyperlink') and cell.hyperlink is not None
                link_target = cell.hyperlink.target if has_link else None
                print(f"  列{col_idx}({headers[col_idx]}): {str(cell.value)[:50]}... | 超链接: {has_link} | target: {link_target}")
    
    wb.close()

print("\n\n读取完成")
